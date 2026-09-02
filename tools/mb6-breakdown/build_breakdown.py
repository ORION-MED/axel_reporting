"""Разбор обязательных видов СЭМД у одной медорганизации — форма для методолога.

Зачем. На ВКС 24.08.2026 методолог про ГБУ «Межрайонная больница № 6»:
«50 видов из 145 регистрирует — да, это правда, так оно и есть. 81 вид для неё
является обязательным — ну, вот это уже неправда. Поменьше, поменьше».

После применения перечней МО их стало 82: перечни сузили состав у других
организаций, а у МБ № 6 виды, лежавшие в «не определено», перешли
в обязательные. Спорить с методологом цифрой бессмысленно — нужен список,
по которому она скажет, какие виды лишние и почему.

Форма устроена так, чтобы отвечать было можно группами. Виды сгруппированы
по условию применимости: если у больницы нет лицензии на медосмотры, снимаются
сразу девять видов, и отмечать каждый по отдельности не нужно.

Данные берутся из TSV-выгрузки стенда — запрос в README.

Запуск:
    python build_breakdown.py mb6.tsv "ГБУ «Межрайонная больница № 6»" out.xlsx
"""

from __future__ import annotations

import csv
import io
import re
import sys
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DECISIONS = ['оставить', 'убрать']

# Строка формы без условия вовсе: вид становится обязателен всем 37 МО.
# Подпись называет это прямо — иначе группа читается как «условие потерялось
# при разборе», хотя потерялось оно в самой форме.
UNCONDITIONAL = 'условия в форме нет — вид обязателен всем 37 МО'

HEADER_FILL = PatternFill('solid', fgColor='DDE5F0')
GROUP_FILL = PatternFill('solid', fgColor='F2F2F2')
ANSWER_FILL = PatternFill('solid', fgColor='FFF6DA')


def read_rows(path: str) -> list[dict[str, str]]:
    with io.open(path, encoding='utf-8') as handle:
        rows = list(csv.DictReader(handle, delimiter='\t'))
    return [row for row in rows if row.get('code') and row.get('status')]


def condition_of(reason: str) -> str:
    """Условие применимости из основания.

    Основание выглядит как «Обязателен по матрице применимости (строка 28:
    Амбулаторный · наличие прикрепленного населения)». Нужна часть после
    двоеточия: именно она объясняет, почему вид попал к этой МО, и именно
    её методолог будет оспаривать.

    Строк в основании бывает несколько («строка 15: Амбулаторный; строка 16:
    Стационарный») — берётся первая: группировка нужна для укрупнения, а не
    для точности до строки формы.
    """
    match = re.search(r'\(строка[^:)]*:?\s*(.*?)\)\.?\s*$', reason or '')
    if not match:
        return UNCONDITIONAL
    body = match.group(1).strip()
    if not body:
        return UNCONDITIONAL
    return body.split(';')[0].strip()


def build(rows: list[dict[str, str]], organization: str, out_path: str) -> None:
    required = [row for row in rows if row['status'] == 'required']
    unknown = [row for row in rows if row['status'] == 'unknown']

    by_condition: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in required:
        by_condition[condition_of(row['reason'])].append(row)

    # Условия — по убыванию числа видов: снимать группой выгоднее с большой.
    order = [c for c, _ in Counter(
        {c: len(items) for c, items in by_condition.items()}
    ).most_common()]

    workbook = Workbook()
    write_types_sheet(workbook.active, organization, by_condition, order, unknown)
    write_summary_sheet(workbook.create_sheet('Свод по условиям'),
                        by_condition, order)
    workbook.save(out_path)


def write_types_sheet(sheet, organization, by_condition, order, unknown) -> None:
    sheet.title = 'Обязательные виды'
    total = sum(len(items) for items in by_condition.values())
    registered = sum(
        1 for items in by_condition.values()
        for row in items if int(row['docs'] or 0) > 0
    )

    sheet.append([f'{organization}: обязательные виды СЭМД по матрице применимости'])
    sheet['A1'].font = Font(bold=True, size=13)
    sheet.append([
        f'Всего обязательных: {total}. Из них регистрируются: {registered}, '
        f'не регистрируются: {total - registered}.'
    ])
    sheet.append([
        'Виды сгруппированы по условию применимости — снимать можно сразу группой.'
    ])
    sheet.append([])

    headers = [
        'Код вида', 'Наименование вида', 'Условие применимости',
        'Регистрирует', 'Документов', 'Решение', 'Комментарий',
    ]
    sheet.append(headers)
    header_row = sheet.max_row
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(header_row, column)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    for condition in order:
        items = by_condition[condition]
        sheet.append([f'{condition} — {len(items)} видов'])
        group_row = sheet.max_row
        sheet.cell(group_row, 1).font = Font(bold=True)
        for column in range(1, len(headers) + 1):
            sheet.cell(group_row, column).fill = GROUP_FILL

        for row in sorted(items, key=lambda item: int(item['code'])):
            docs = int(row['docs'] or 0)
            sheet.append([
                row['code'],
                row['name'],
                condition,
                'да' if docs else 'нет',
                docs,
                '',
                '',
            ])
            sheet.cell(sheet.max_row, 6).fill = ANSWER_FILL
            sheet.cell(sheet.max_row, 7).fill = ANSWER_FILL

    if unknown:
        sheet.append([])
        sheet.append([
            f'Справочно: ещё {len(unknown)} видов в состоянии «не определено» — '
            'условие формы не удалось разобрать. В число обязательных они не входят.'
        ])
        sheet.cell(sheet.max_row, 1).font = Font(italic=True)
        for row in sorted(unknown, key=lambda item: int(item['code'])):
            sheet.append([row['code'], row['name'], 'не определено'])

    first_data_row = header_row + 1
    validation = DataValidation(
        type='list',
        formula1='"' + ','.join(DECISIONS) + '"',
        allow_blank=True,
        showErrorMessage=True,
    )
    sheet.add_data_validation(validation)
    validation.add(f'F{first_data_row}:F{sheet.max_row}')

    widths = [10, 62, 46, 12, 12, 14, 40]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = f'A{first_data_row}'


def write_summary_sheet(sheet, by_condition, order) -> None:
    """Свод: сколько видов держит каждое условие и сколько из них живые.

    Он и есть главный лист для разговора. Если условие «наличие лицензии
    на медосмотры» даёт девять видов и ни один не регистрируется — это первый
    кандидат в лишние, и виден он здесь, а не в списке из восьмидесяти двух строк.
    """
    sheet.append(['Условие применимости', 'Видов', 'Регистрируются', 'Не регистрируются'])
    for column in range(1, 5):
        cell = sheet.cell(1, column)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for condition in order:
        items = by_condition[condition]
        registered = sum(1 for row in items if int(row['docs'] or 0) > 0)
        sheet.append([condition, len(items), registered, len(items) - registered])

    sheet.column_dimensions['A'].width = 76
    for column in 'BCD':
        sheet.column_dimensions[column].width = 16
    sheet.freeze_panes = 'A2'


def main() -> None:
    if len(sys.argv) != 4:
        print(__doc__)
        raise SystemExit(2)
    source, organization, out_path = sys.argv[1:4]
    rows = read_rows(source)
    build(rows, organization, out_path)
    print(f'Готово: {out_path}, строк исходных {len(rows)}')


if __name__ == '__main__':
    main()
