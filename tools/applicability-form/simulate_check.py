# -*- coding: utf-8 -*-
"""Повторяет формулу колонки «Проверка строки» на Python: Excel её посчитает
только при открытии, а знать заранее, что увидит методолог, нужно сейчас."""
import collections
import openpyxl

import sys

P = ("D:/dev/axel_actual/axel_server_ready/axel_server_ready/AXEL_контекст/24.08/"
     "8_Матрица_применимости_145_видов_с_перечнями_МО_2026-08-24.xlsx")
if len(sys.argv) == 2:
    P = sys.argv[1]

wb = openpyxl.load_workbook(P, data_only=True)
ws = wb["Форма условий"]
names = [str(c[0].value).replace(" ", "").upper()
         for c in wb["Справочник ФРМО"].iter_rows(min_row=2, max_row=38,
                                                  min_col=23, max_col=23)]


def bad(cell_value):
    if not cell_value:
        return False
    packed = "," + str(cell_value).replace(" ", "").upper() + ","
    found = sum(1 for n in names if ("," + n + ",") in packed)
    expected = str(cell_value).count(",") + 1
    return found < expected


def verdict(r):
    g = lambda c: (str(ws.cell(r, c).value).strip() if ws.cell(r, c).value is not None else "")
    D, E, F, G, H, I = g(4), g(5), g(6), g(7), g(8), g(9)
    if D == "":
        return "НЕ ЗАПОЛНЕНО"
    if H and I:
        return "ОШИБКА: обе колонки"
    if bad(H):
        return "ОШИБКА: имя не из справочника (H)"
    if bad(I):
        return "ОШИБКА: имя не из справочника (I)"
    if "если МО" in G and not H and not I:
        return "НУЖНО РЕШЕНИЕ: перечень не распознан"
    if D == "условно":
        return "НУЖНО РЕШЕНИЕ: условно"
    if D == "не определено":
        return "НУЖНО РЕШЕНИЕ"
    if not E and not F and not G and not H and not I:
        return "ГОТОВО: ДЛЯ ВСЕХ"
    return "ГОТОВО"


counts = collections.Counter()
problems = []
for r in range(5, 583):
    if not ws.cell(r, 2).value:
        continue
    v = verdict(r)
    counts[v] += 1
    if not v.startswith("ГОТОВО"):
        problems.append((r, ws.cell(r, 2).value, ws.cell(r, 4).value, v))

print("Что покажет «Проверка строки»:")
for k, v in counts.most_common():
    print("  %-40s %d" % (k, v))
print("\nСтроки, требующие внимания методолога (%d):" % len(problems))
for r, code, dec, v in problems:
    print("  стр %3d | вид %-5s | %-14s | %s" % (r, code, dec, v))
