# -*- coding: utf-8 -*-
"""
Пересборка формы условий обязательности: две колонки перечней МО с выпадающим
списком кратких наименований. Решение ВКС 24.08.2026 (В-04).

Колонки вставляются после «Дополнительного условия», всё правее сдвигается на две.
Формулы «Проверки строки» ссылаются только на A..G — сдвиг их не ломает.
"""
import re
import sys
from copy import copy
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter, column_index_from_string

BASE = "D:/dev/axel_actual/axel_server_ready/axel_server_ready/AXEL_контекст/"
SRC_MATRIX = BASE + "24.08/8_Матрица_применимости_145_видов_от_13.08.xlsx"
SRC_DIR = BASE + "24.08/7_МО_Курганской_области_справочник_признаков.xlsx"
OUT = BASE + "24.08/8_Матрица_применимости_145_видов_с_перечнями_МО_2026-08-24.xlsx"

# Пути по умолчанию — редакция от 24.08.2026. Для следующей редакции
# передайте три аргумента: матрица, справочник признаков МО, результат.
if len(sys.argv) == 4:
    SRC_MATRIX, SRC_DIR, OUT = sys.argv[1], sys.argv[2], sys.argv[3]

SHEET = "Форма условий"
REF = "Справочник ФРМО"
LAST_ROW = 582
INS_AT = 8
N_NEW = 2
NAMES_COL = 23

wbd = openpyxl.load_workbook(SRC_DIR, data_only=True)
wsd = wbd["Лист1"]
short_names = []
for row in wsd.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    name = str(row[3]).strip() if row[3] else ""
    if name:
        short_names.append(name)
assert len(short_names) == 37, len(short_names)


def norm(s):
    s = str(s).upper().replace("Ё", "Е")
    s = re.sub(r'["«»№]', "", s)
    s = re.sub(r"^\s*(ГБУ|ГКУ|МАУЗ|ФГБУ|ГАУЗ|ОБУЗ|БУ|АО)\s+", "", s)
    return re.sub(r"[\s\-]+", "", s).strip()


by_norm = {norm(n): n for n in short_names}
SYNONYMS = {
    # Написание методолога -> краткое имя из справочника признаков МО.
    # Каждая пара — наше допущение, вынесено в сопроводительную записку на проверку.
    "КОПБ": "КОПНБ",              # иного кандидата среди 37 МО нет
    "БЮРОСМЭ": "КОБСМЭ",
    "КУРГАНФАРМАЦИЯ": "АО КФ",
    "ГСП": "МАУЗ ГСП",
    "КОЦМП": "КОЦОЗМП",           # центр медицинской профилактики; КОЦПБС — это СПИД-центр
}


def resolve(token):
    k = norm(token)
    if k in by_norm:
        return by_norm[k]
    return SYNONYMS.get(k)


PREFIX = re.compile(r"^\s*если\s+мо(?=\s|[-–—:]|$)\s*", re.I | re.U)
NEG = re.compile(r"^(?:[-–—:]\s*)?не\s+", re.I | re.U)


def parse_condition(text):
    t = re.sub(r"\s+", " ", str(text or "")).strip()
    if not PREFIX.match(t):
        return [], False
    rest = PREFIX.sub("", t)
    excluded = bool(NEG.match(rest))
    if excluded:
        rest = NEG.sub("", rest)
    rest = re.sub(r"^[-–—:]\s*", "", rest)
    parts = [p.strip(" .-–—") for p in re.split(r"[,;/]|\sи\s", rest)]
    return [p for p in parts if p], excluded


wb = openpyxl.load_workbook(SRC_MATRIX)
old = wb[SHEET]
sheet_index = wb.sheetnames.index(SHEET)

MAX_COL = 26
cells = {}
for r in range(1, LAST_ROW + 1):
    for c in range(1, MAX_COL + 1):
        cell = old.cell(row=r, column=c)
        cells[(r, c)] = (cell.value, copy(cell._style))


def newcol(c):
    return c if c < INS_AT else c + N_NEW


new = wb.create_sheet("__tmp__", sheet_index)
for (r, c), (value, style) in cells.items():
    tgt = new.cell(row=r, column=newcol(c))
    tgt.value = value
    tgt._style = style

# Заголовки намеренно нейтральные, а не «Обязателен только этим МО». Перечень
# применяется к строке целиком, а решение строки может быть и «не обязателен»:
# у вида 85 стоит «не обязателен» с перечнем из шести узкоспециализированных МО,
# и заголовок со словом «обязателен» прочитался бы ровно наоборот.
header_style = copy(old.cell(row=4, column=7)._style)
for col, title in (
    (INS_AT, "Только эти МО"),
    (INS_AT + 1, "Все МО, кроме этих"),
):
    h = new.cell(row=4, column=col)
    h.value = title
    h._style = header_style

body_style = copy(old.cell(row=5, column=7)._style)
filled_rows = 0
unresolved = {}
for r in range(5, LAST_ROW + 1):
    for col in (INS_AT, INS_AT + 1):
        new.cell(row=r, column=col)._style = copy(body_style)
    code = cells[(r, 2)][0]
    if code in (None, ""):
        continue
    names, excluded = parse_condition(cells[(r, 7)][0])
    if not names:
        continue
    resolved, missing = [], []
    for token in names:
        hit = resolve(token)
        if hit:
            resolved.append(hit)
        else:
            missing.append(token)
    for token in missing:
        unresolved.setdefault(token, set()).add(str(code))
    if missing or not resolved:
        continue
    new.cell(row=r, column=INS_AT + 1 if excluded else INS_AT).value = ", ".join(resolved)
    filled_rows += 1

# Вид 47: условие «прикреплённое население» — это не перечень МО, а признак,
# поэтому оно живёт в «Дополнительном условии», а не в новых колонках. В файле
# методолога от 24.08 оно пропало; без него вид становится обязателен 33 МО
# вместо 15, и ГКУ «КОПНБ» откатывается со 100 % на 13 из 14. Ответ на В-02
# получен дважды — письменно 21.08 и голосом на ВКС 24.08, — вносим и выносим
# в сопроводительную записку на подтверждение.
V47_CONDITION = "Обязателен для МО с прикреплённым населением"
v47_rows = []
for r in range(5, LAST_ROW + 1):
    if str(cells[(r, 2)][0] or "").strip() == "47" and not str(cells[(r, 7)][0] or "").strip():
        new.cell(row=r, column=7).value = V47_CONDITION
        v47_rows.append(r)

for letter, dim in old.column_dimensions.items():
    idx = column_index_from_string(letter)
    tgt = get_column_letter(newcol(idx))
    if dim.width:
        new.column_dimensions[tgt].width = dim.width
new.column_dimensions[get_column_letter(INS_AT)].width = 30.0
new.column_dimensions[get_column_letter(INS_AT + 1)].width = 30.0

for idx, dim in old.row_dimensions.items():
    if dim.height:
        new.row_dimensions[idx].height = dim.height

new.freeze_panes = old.freeze_panes
new.sheet_view.showGridLines = old.sheet_view.showGridLines

for merged in ("A1:K1", "A2:K2", "A3:P3"):
    new.merge_cells(merged)

# Подпись под шапкой должна назвать новые колонки: иначе методолог по привычке
# продолжит писать перечень МО текстом в «Дополнительном условии».
new.cell(row=3, column=1).value = (
    "Пустой «Тип подразделения» или «Вид подразделения» = любой. "
    "В «Дополнительном условии» опишите проверяемое условие: прикреплённое население, "
    "лицензия, профиль помощи. Перечень конкретных МО — в колонках «Только эти МО» "
    "и «Все МО, кроме этих», выбором из выпадающего списка."
)

# Решение ВКС 24.08.2026: «зачем нам тогда решение "условно", если мы его заменим».
# Убираем «условно» из выпадающего списка «Решение», чтобы оно не появлялось
# в новых строках. В девятнадцати существующих строках значение остаётся —
# Excel проверяет ввод, а не то, что уже записано, — и «Проверка строки»
# по каждой из них требует замены.
wsr = wb[REF]
wsr.cell(row=5, column=21).value = "не определено"
wsr.cell(row=6, column=21).value = None
wb.defined_names["RULE_STATUSES"] = DefinedName(
    "RULE_STATUSES", attr_text="'%s'!$U$3:$U$5" % REF
)

wsr.cell(row=1, column=NAMES_COL).value = "Краткие наименования МО (для колонок H и I формы)"
wsr.cell(row=1, column=NAMES_COL).font = Font(bold=True)
for i, name in enumerate(short_names):
    wsr.cell(row=2 + i, column=NAMES_COL).value = name
wsr.column_dimensions[get_column_letter(NAMES_COL)].width = 32
last = 1 + len(short_names)
col_letter = get_column_letter(NAMES_COL)
wb.defined_names["MO_SHORT_NAMES"] = DefinedName(
    "MO_SHORT_NAMES",
    attr_text="'%s'!$%s$2:$%s$%d" % (REF, col_letter, col_letter, last),
)

for dv in old.data_validations.dataValidation:
    copy_dv = DataValidation(
        type=dv.type,
        formula1=dv.formula1,
        allow_blank=dv.allowBlank,
        showDropDown=dv.showDropDown,
        showErrorMessage=dv.showErrorMessage,
    )
    new.add_data_validation(copy_dv)
    copy_dv.sqref = dv.sqref

mo_dv = DataValidation(
    type="list", formula1="MO_SHORT_NAMES", allow_blank=True, showErrorMessage=False
)
mo_dv.promptTitle = "Перечень МО"
mo_dv.prompt = (
    "Выберите МО из списка. Несколько МО — через запятую. "
    "Наименования берутся из справочника признаков МО."
)
mo_dv.showInputMessage = True
new.add_data_validation(mo_dv)
mo_dv.sqref = "H5:I%d" % LAST_ROW

CHECK_COL = newcol(9)
CH = get_column_letter(CHECK_COL)
H = get_column_letter(INS_AT)
I = get_column_letter(INS_AT + 1)


def bad_list(col, row):
    packed_cell = '","&SUBSTITUTE($%s%d," ","")&","' % (col, row)
    packed_dir = '","&SUBSTITUTE(MO_SHORT_NAMES," ","")&","'
    found = "SUMPRODUCT(--ISNUMBER(SEARCH(%s,%s)))" % (packed_dir, packed_cell)
    expected = 'LEN(TRIM($%s%d))-LEN(SUBSTITUTE(TRIM($%s%d),",",""))+1' % (col, row, col, row)
    return 'AND($%s%d<>"",%s<%s)' % (col, row, found, expected)


def check_formula(row):
    return (
        '=IF($D{r}="","НЕ ЗАПОЛНЕНО",'
        'IF(AND(${H}{r}<>"",${I}{r}<>""),"ОШИБКА: заполнены обе колонки перечня — оставьте одну",'
        'IF({badH},"ОШИБКА: в «Только эти МО» наименование не из справочника",'
        'IF({badI},"ОШИБКА: в «Все МО, кроме этих» наименование не из справочника",'
        # Нераспознанный перечень проверяется РАНЬШЕ «условно»: у видов 90 и 142
        # стоит и то и другое, и при обратном порядке настоящий блокер — перечень —
        # всплыл бы только вторым заходом, после замены слова «условно».
        'IF(AND(ISNUMBER(SEARCH("если МО",$G{r})),${H}{r}="",${I}{r}=""),'
        '"НУЖНО РЕШЕНИЕ: перечень МО из условия не распознан — впишите его в колонку H или I",'
        'IF($D{r}="условно","НУЖНО РЕШЕНИЕ: замените «условно» на «обязателен» и укажите перечень МО",'
        'IF($D{r}="не определено","НУЖНО РЕШЕНИЕ",'
        'IF(AND($E{r}="",$F{r}="",$G{r}="",${H}{r}="",${I}{r}=""),"ГОТОВО: ДЛЯ ВСЕХ","ГОТОВО"))))))))'
    ).format(r=row, H=H, I=I, badH=bad_list(H, row), badI=bad_list(I, row))


for r in range(5, LAST_ROW + 1):
    new.cell(row=r, column=CHECK_COL).value = check_formula(r)

GREEN, YELLOW, RED = "FFD9EAD3", "FFFFF2CC", "FFF4CCCC"
rng = "%s5:%s%d" % (CH, CH, LAST_ROW)
for keyword, color in (
    ("ГОТОВО", GREEN),
    ("НУЖНО", YELLOW),
    ("НЕ ЗАПОЛНЕНО", RED),
    ("ОШИБКА", RED),
):
    new.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=['ISNUMBER(SEARCH("%s",%s5))' % (keyword, CH)],
            fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
        ),
    )

del wb[SHEET]
new.title = SHEET
wb.move_sheet(SHEET, offset=sheet_index - wb.sheetnames.index(SHEET))

wc = wb["Контроль заполнения"]
for r in range(13, wc.max_row + 1):
    cell = wc.cell(row=r, column=7)
    if isinstance(cell.value, str) and cell.value.startswith("="):
        cell.value = re.sub(r"!I(\d+)$", lambda m: "!%s%s" % (CH, m.group(1)), cell.value)

wc["A10"] = "Строк с перечнем МО"
wc["B10"] = "=COUNTA('%s'!$H$5:$H$231)+COUNTA('%s'!$I$5:$I$231)" % (SHEET, SHEET)
wc["A11"] = "Ошибок и нерешённых строк"
wc["B11"] = (
    '=SUMPRODUCT(--ISNUMBER(SEARCH("ОШИБКА",\'%s\'!$%s$5:$%s$231)))'
    '+SUMPRODUCT(--ISNUMBER(SEARCH("НУЖНО",\'%s\'!$%s$5:$%s$231)))'
) % (SHEET, CH, CH, SHEET, CH, CH)
for coord in ("A10", "A11"):
    wc[coord].font = copy(wc["A9"].font)
    wc[coord].alignment = copy(wc["A9"].alignment)
for coord in ("B10", "B11"):
    wc[coord].font = copy(wc["B9"].font)

wh = wb["Как заполнять"]
wh["A21"] = (
    "условно — БОЛЬШЕ НЕ ИСПОЛЬЗУЕТСЯ (решение ВКС 24.08.2026). Замените на «обязателен» "
    "и укажите перечень МО в колонке H или I."
)
bold = copy(wh["A4"].font)
plain = copy(wh["A5"].font)
HOWTO = [
    ("ЧТО ДОБАВЛЕНО 24.08.2026 — ДВЕ КОЛОНКИ ПЕРЕЧНЕЙ МО", True),
    (
        "Решение ВКС 24.08.2026 по вопросу В-04: перечень медорганизаций больше не пишется "
        "текстом в «Дополнительном условии», а выбирается в двух новых колонках.",
        False,
    ),
    ("", False),
    ("H — «Только эти МО»", True),
    (
        "Строка действует ИСКЛЮЧИТЕЛЬНО на перечисленные медорганизации. "
        "Пример: вид 371 «Протокол консилиума врачей онкологического», решение «обязателен», "
        "H = КООД — значит обязателен только онкодиспансеру, остальным не обязателен.",
        False,
    ),
    ("I — «Все МО, кроме этих»", True),
    (
        "Строка действует на всех, кто прошёл остальные условия, кроме перечисленных. "
        "Пример: вид 86 «Рецепт 107-1/у», решение «обязателен», I = МАУЗ ГСП — "
        "обязателен всем амбулаторным, кроме стоматологической поликлиники.",
        False,
    ),
    (
        "ВАЖНО: перечень относится к строке целиком, а не к слову «обязателен». Если решение "
        "строки «не обязателен», то H означает «не обязателен только этим МО». Так стоит "
        "у вида 85: решение «не обязателен», H — шесть узкоспециализированных МО.",
        False,
    ),
    ("", False),
    ("Как заполнять колонки", True),
    (
        "Щёлкните по ячейке — появится выпадающий список кратких наименований всех 37 МО. "
        "Несколько МО перечисляются через запятую: «КОПАБ, КОБСМЭ». Список подставляет одно "
        "наименование; второе и последующие допишите вручную через запятую.",
        False,
    ),
    (
        "Заполнять нужно ТОЛЬКО ОДНУ из двух колонок. Если заполнены обе — «Проверка строки» "
        "покажет ошибку.",
        False,
    ),
    (
        "Наименования берутся из колонки «краткое наименование для отображения в сервисе» "
        "справочника признаков МО. Полный перечень — на листе «Справочник ФРМО», колонка W.",
        False,
    ),
    ("", False),
    ("Что уже заполнено за Вас", True),
    (
        "Перечни, которые удалось разобрать из «Дополнительного условия», перенесены "
        "в колонки автоматически — 25 строк. Исходный текст условия оставлен на месте "
        "для сверки; при импорте выигрывает колонка.",
        False,
    ),
    (
        "Три наименования разобрать не удалось, их нужно указать вручную: «Санаторий» "
        "(виды 48, 50, 357), «Диспансер» (вид 90), «Психоневрологический диспансер» (вид 142). "
        "В «Проверке строки» они помечены жёлтым.",
        False,
    ),
    ("", False),
    ("Проверка строки — что означают пометки", True),
    ("ГОТОВО — строка заполнена, вопросов нет.", False),
    ("НУЖНО РЕШЕНИЕ — осталось «условно», «не определено» или неразобранный перечень МО.", False),
    ("ОШИБКА — заполнены обе колонки перечня либо указано наименование не из справочника.", False),
    (
        "Сводка по всей форме — на листе «Контроль заполнения», строки «Строк с перечнем МО» "
        "и «Ошибок и нерешённых строк». Цель — ноль во второй строке.",
        False,
    ),
]
row = 58
for text, is_bold in HOWTO:
    cell = wh.cell(row=row, column=1)
    cell.value = text
    cell.font = copy(bold if is_bold else plain)
    row += 1

wb.save(OUT)

print("Сохранено:", OUT)
print("Кратких наименований в справочнике:", len(short_names))
print("Строк с предзаполненным перечнем:", filled_rows)
print("Вид 47 — условие возвращено в строках:", v47_rows)
print("Не сопоставилось со справочником (оставлено методологу):")
for token, codes in sorted(unresolved.items()):
    print("  <%s> - виды %s" % (token, ", ".join(sorted(codes, key=int))))
