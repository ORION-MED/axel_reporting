# -*- coding: utf-8 -*-
"""
Форма отчёта из региональной ГИС: закрытые случаи по месяцам.

ВКС 24.08.2026. Николай Ермаков: «третья колонка — законченные случаи по данным
ГИС… он может быть краснее красного». Методолог: «регион может заказывать отчёт
вендору — мы даём форму, по которой нам было бы удобно его подгружать,
техподдержка МИАЦ запрашивает по этой форме отчёт и загружает в программу».

Ключевое требование к форме — сопоставимость. Виды случая заданы не «как удобно»,
а один в один по листам ТПГГ, которые уже стоят в знаменателях показателей
(`semd-volume-ratio.config.ts`). Иначе сравнивать «план — факт СЭМД — закрытые
случаи» будет не с чем: три числа окажутся про разные совокупности.

Запуск:
    python tools/gis-report-form/build_gis_form.py [справочник_МО.xlsx] [результат.xlsx]
"""
import sys

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

BASE = "D:/dev/axel_actual/axel_server_ready/axel_server_ready/AXEL_контекст/"
SRC_DIR = BASE + "24.08/7_МО_Курганской_области_справочник_признаков.xlsx"
OUT = BASE + "24.08/AXEL_форма_отчёта_из_ГИС_закрытые_случаи_2026-08-24.xlsx"
if len(sys.argv) == 3:
    SRC_DIR, OUT = sys.argv[1], sys.argv[2]

YEAR = 2026
MONTHS = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

# Виды случая = листы ТПГГ из знаменателей четырёх показателей-долей.
# Код листа совпадает с `sheet_code` в `reporting_tpgg_plan_values` — по нему
# импортёр и сведёт закрытые случаи с планом.
# (код листа, наименование вида случая, показатель)
CASE_TYPES = [
    ("1", "Скорая медицинская помощь — вызовы", "6.1.3.2.11"),
    ("2", "Обращения по заболеваниям (амбулаторно)", "6.1.3.2.8"),
    ("3", "Посещения с иными целями", "6.1.3.2.8"),
    ("4", "Неотложная помощь", "6.1.3.2.8"),
    ("3.2", "Диспансеризация определённых групп взрослого населения", "6.1.3.2.9"),
    ("3.3", "Углублённая диспансеризация", "6.1.3.2.9"),
    ("3.4", "Диспансеризация репродуктивного здоровья женщин", "6.1.3.2.9"),
    ("3.5", "Диспансеризация репродуктивного здоровья мужчин", "6.1.3.2.9"),
    ("3.6", "Диспансеризация детей-сирот", "6.1.3.2.9"),
    ("3.7", "Диспансеризация детей под опекой", "6.1.3.2.9"),
    ("3.8", "Профилактические медицинские осмотры взрослых", "6.1.3.2.9"),
    ("3.9", "Профилактические медицинские осмотры несовершеннолетних", "6.1.3.2.9"),
    ("5", "Круглосуточный стационар", "6.1.3.2.10"),
    ("6", "Высокотехнологичная медицинская помощь", "6.1.3.2.10"),
    ("7", "Медицинская реабилитация в круглосуточном стационаре", "6.1.3.2.10"),
    ("8", "Дневные стационары", "6.1.3.2.10"),
    ("9", "Медицинская реабилитация в дневном стационаре", "6.1.3.2.10"),
]

INDICATOR_TITLES = {
    "6.1.3.2.8": "Эпикриз по законченному случаю амбулаторный",
    "6.1.3.2.9": "Результаты профилактического медицинского осмотра (диспансеризации)",
    "6.1.3.2.10": "Эпикриз в стационаре выписной / выписной эпикриз из родильного дома",
    "6.1.3.2.11": "Карта вызова скорой медицинской помощи",
}

HEADER_FILL = PatternFill("solid", start_color="FFDCE6F1", end_color="FFDCE6F1")
INPUT_FILL = PatternFill("solid", start_color="FFFFF9E0", end_color="FFFFF9E0")
TITLE_FONT = Font(bold=True, size=13)
HEAD_FONT = Font(bold=True)
THIN = Side(style="thin", color="FFB0B8C4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def read_organizations(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Лист1"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append({
            "oid": str(row[0]).strip(),
            "name": str(row[2] or "").strip(),
            "short": str(row[3] or "").strip(),
        })
    return rows


def put(ws, row, col, value, font=None, fill=None, align=None, border=True, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def build_howto(ws, organizations):
    ws.column_dimensions["A"].width = 118
    blocks = [
        ("Отчёт из региональной ГИС: закрытые случаи по месяцам", "title"),
        ("Назначение: сопоставить план по территориальной программе, количество "
         "зарегистрированных в РЭМД СЭМД и фактическое количество закрытых случаев "
         "в региональной ГИС. Разрыв между закрытыми случаями и СЭМД показывает, "
         "формируются ли документы вовремя или накапливаются к концу года.", ""),
        ("", ""),
        ("Кто заполняет", "head"),
        ("Отчёт формируется на стороне региональной ГИС — силами вендора по заявке МИАЦ "
         "либо выгрузкой штатными средствами. Данные обезличенные: только количества, "
         "без списков пациентов и без персональных данных.", ""),
        ("", ""),
        ("Что считать закрытым случаем", "head"),
        ("Случай, закрытый в региональной ГИС за отчётный месяц, — независимо от того, "
         "сформирован ли по нему СЭМД и отправлен ли он на регистрацию. Именно это "
         "и сравнивается: сколько случаев закрыто против того, сколько документов "
         "фактически ушло в РЭМД.", ""),
        ("Случай учитывается в том месяце, в котором он закрыт, а не в котором открыт.", ""),
        ("", ""),
        ("Листы формы", "head"),
        ("«Закрытые случаи» — обязательный. Одна строка на медорганизацию и вид случая, "
         "по колонке на каждый месяц.", ""),
        ("«По подразделениям» — необязательный. Та же таблица, но с разбивкой по "
         "структурным подразделениям и зданиям. Если ГИС отдаёт такую детализацию — "
         "заполните: она позволит показать положение дел не только по больнице целиком, "
         "но и по отдельным подразделениям.", ""),
        ("", ""),
        ("Как заполнять", "head"),
        ("В колонках месяцев — количество закрытых случаев за этот месяц, а не "
         "нарастающим итогом. Накопительный итог посчитаем сами.", ""),
        ("Заполняйте месяцы по текущий включительно; будущие оставьте пустыми.", ""),
        ("Если медорганизация такую помощь не оказывает — оставьте строку пустой. "
         "Это не ошибка, строки заведены на все сочетания заранее.", ""),
        ("Ноль ставьте только тогда, когда помощь оказывается, но за месяц "
         "не закрыто ни одного случая. Пустая ячейка и ноль — разные вещи.", ""),
        ("", ""),
        ("Виды случая — почему именно такие", "head"),
        ("Перечень видов случая повторяет разделы территориальной программы, по которым "
         "уже считаются знаменатели показателей. Коды в колонке «Код вида случая» — "
         "это номера листов территориальной программы. Менять и дополнять перечень "
         "не нужно: при другом составе сравнение с планом станет некорректным.", ""),
        ("Расшифровка — на листе «Виды случая», там же указано, в какой показатель "
         "входит каждый вид.", ""),
        ("", ""),
        ("Пример", "head"),
        ("Межрайонная больница № 6, вид случая 5 «Круглосуточный стационар»: "
         "в январе закрыто 812 случаев, в феврале 764. В строке этой больницы "
         "и этого вида случая в колонке «Январь» ставится 812, в «Феврале» — 764.", ""),
        ("", ""),
        ("Проверка", "head"),
        ("Лист «Контроль заполнения» пересчитывается сам: сколько строк заполнено, "
         "сколько случаев набралось по каждому показателю и есть ли ячейки "
         "с нечисловыми значениями. Перед отправкой убедитесь, что строка "
         "«Ячеек с нечисловым значением» показывает ноль.", ""),
        ("", ""),
        ("Справочник медорганизаций", "head"),
        ("На листе «Справочник МО» перечислены все %d медорганизаций с их OID по ФРМО. "
         "OID — ключ, по которому отчёт связывается с остальными данными сервиса; "
         "менять его нельзя." % len(organizations), ""),
    ]
    for index, (text, kind) in enumerate(blocks, start=1):
        cell = ws.cell(row=index, column=1)
        cell.value = text
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "title":
            cell.font = TITLE_FONT
        elif kind == "head":
            cell.font = HEAD_FONT


def build_data_sheet(ws, organizations, by_subdivision):
    """Обязательный лист по МО либо необязательный с разбивкой по подразделениям."""
    if by_subdivision:
        head = ["OID МО", "Наименование МО", "OID структурного подразделения",
                "Название подразделения", "ID здания", "Адрес здания",
                "Код вида случая", "Вид случая"]
        widths = [42, 34, 42, 30, 12, 34, 15, 40]
    else:
        head = ["№", "OID МО", "Наименование МО", "Код вида случая", "Вид случая", "Показатель"]
        widths = [6, 42, 34, 15, 46, 14]

    title = ("Закрытые случаи в региональной ГИС за %d год — по подразделениям"
             if by_subdivision else
             "Закрытые случаи в региональной ГИС за %d год") % YEAR
    put(ws, 1, 1, title, font=TITLE_FONT, border=False)
    subtitle = (
        "Необязательный лист. Заполняется, если ГИС отдаёт разбивку по подразделениям. "
        "Строки не заведены заранее — добавляйте по мере необходимости."
        if by_subdivision else
        "Количество случаев, закрытых за месяц, — не нарастающим итогом. "
        "Строки, где помощь не оказывается, оставьте пустыми."
    )
    put(ws, 2, 1, subtitle, border=False)
    put(ws, 3, 1, "Жёлтые ячейки — для заполнения. Остальное менять не нужно.", border=False)

    for offset, name in enumerate(head):
        col = offset + 1
        put(ws, 4, col, name, font=HEAD_FONT, fill=HEADER_FILL, align=WRAP_CENTER)
        ws.column_dimensions[get_column_letter(col)].width = widths[offset]
    month_start = len(head) + 1
    for offset, month in enumerate(MONTHS):
        col = month_start + offset
        put(ws, 4, col, month, font=HEAD_FONT, fill=HEADER_FILL, align=WRAP_CENTER)
        ws.column_dimensions[get_column_letter(col)].width = 11
    total_col = month_start + len(MONTHS)
    put(ws, 4, total_col, "Итого за год", font=HEAD_FONT, fill=HEADER_FILL, align=WRAP_CENTER)
    ws.column_dimensions[get_column_letter(total_col)].width = 14
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = ws.cell(row=5, column=month_start)

    last_row = 4
    if not by_subdivision:
        row = 5
        number = 0
        for organization in organizations:
            for code, case_name, indicator in CASE_TYPES:
                number += 1
                put(ws, row, 1, number, align=WRAP_CENTER)
                put(ws, row, 2, organization["oid"], align=WRAP)
                put(ws, row, 3, organization["name"], align=WRAP)
                put(ws, row, 4, code, align=WRAP_CENTER)
                put(ws, row, 5, case_name, align=WRAP)
                put(ws, row, 6, indicator, align=WRAP_CENTER)
                for offset in range(len(MONTHS)):
                    put(ws, row, month_start + offset, None,
                        fill=INPUT_FILL, align=WRAP_CENTER, fmt="#,##0")
                first = get_column_letter(month_start)
                last = get_column_letter(month_start + len(MONTHS) - 1)
                put(ws, row, total_col, "=IF(COUNT(%s%d:%s%d)=0,\"\",SUM(%s%d:%s%d))"
                    % (first, row, last, row, first, row, last, row),
                    align=WRAP_CENTER, fmt="#,##0")
                row += 1
        last_row = row - 1
    else:
        # Пустые строки под ручное или машинное заполнение.
        for row in range(5, 405):
            for col in range(1, total_col):
                put(ws, row, col, None, fill=INPUT_FILL, align=WRAP)
            first = get_column_letter(month_start)
            last = get_column_letter(month_start + len(MONTHS) - 1)
            put(ws, row, total_col, "=IF(COUNT(%s%d:%s%d)=0,\"\",SUM(%s%d:%s%d))"
                % (first, row, last, row, first, row, last, row),
                align=WRAP_CENTER, fmt="#,##0")
        last_row = 404

    return {"month_start": month_start, "total_col": total_col, "last_row": last_row}


def build_reference_sheets(wb, organizations):
    ws = wb.create_sheet("Справочник МО")
    put(ws, 1, 1, "Медицинские организации Курганской области", font=TITLE_FONT, border=False)
    put(ws, 2, 1, "Технические данные. Не редактировать — OID связывает отчёт "
                  "с остальными данными сервиса.", border=False)
    for offset, name in enumerate(["OID МО по ФРМО", "Краткое наименование по ФРМО",
                                   "Обозначение в сервисе"]):
        put(ws, 4, offset + 1, name, font=HEAD_FONT, fill=HEADER_FILL, align=WRAP_CENTER)
    for width, letter in zip([42, 40, 22], "ABC"):
        ws.column_dimensions[letter].width = width
    for index, organization in enumerate(organizations):
        row = 5 + index
        put(ws, row, 1, organization["oid"])
        put(ws, row, 2, organization["name"], align=WRAP)
        put(ws, row, 3, organization["short"])
    ws.freeze_panes = "A5"

    cases = wb.create_sheet("Виды случая")
    put(cases, 1, 1, "Виды случая", font=TITLE_FONT, border=False)
    put(cases, 2, 1, "Перечень повторяет разделы территориальной программы, по которым "
                     "считаются знаменатели показателей. Код — номер листа программы.",
        border=False)
    for offset, name in enumerate(["Код вида случая", "Вид случая", "Показатель",
                                   "Наименование показателя"]):
        put(cases, 4, offset + 1, name, font=HEAD_FONT, fill=HEADER_FILL, align=WRAP_CENTER)
    for width, letter in zip([16, 52, 14, 62], "ABCD"):
        cases.column_dimensions[letter].width = width
    for index, (code, case_name, indicator) in enumerate(CASE_TYPES):
        row = 5 + index
        put(cases, row, 1, code, align=WRAP_CENTER)
        put(cases, row, 2, case_name, align=WRAP)
        put(cases, row, 3, indicator, align=WRAP_CENTER)
        put(cases, row, 4, INDICATOR_TITLES[indicator], align=WRAP)
    cases.freeze_panes = "A5"
    return len(organizations), len(CASE_TYPES)


def build_control(wb, layout):
    ws = wb.create_sheet("Контроль заполнения")
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    put(ws, 1, 1, "Контроль заполнения", font=TITLE_FONT, border=False)
    put(ws, 2, 1, "Пересчитывается автоматически при открытии файла.", border=False)

    sheet = "Закрытые случаи"
    first = get_column_letter(layout["month_start"])
    last = get_column_letter(layout["month_start"] + len(MONTHS) - 1)
    total = get_column_letter(layout["total_col"])
    end = layout["last_row"]
    rng = "'%s'!$%s$5:$%s$%d" % (sheet, first, last, end)

    checks = [
        ("Всего строк в форме", "=COUNTA('%s'!$B$5:$B$%d)" % (sheet, end)),
        # Итог по строке — формула, возвращающая "" для пустой строки,
        # поэтому COUNT по колонке итога и есть число заполненных строк.
        ("Заполнено строк", "=COUNT('%s'!$%s$5:$%s$%d)" % (sheet, total, total, end)),
        ("Всего закрытых случаев", "=SUM(%s)" % rng),
        ("Ячеек с нечисловым значением", "=SUMPRODUCT(--ISTEXT(%s))" % rng),
        ("Отрицательных значений", "=SUMPRODUCT(--(%s<0))" % rng),
    ]
    row = 4
    for label, formula in checks:
        put(ws, row, 1, label, font=HEAD_FONT, align=WRAP)
        put(ws, row, 2, formula, align=WRAP_CENTER, fmt="#,##0")
        row += 1

    row += 1
    put(ws, row, 1, "По показателям", font=HEAD_FONT, border=False)
    row += 1
    put(ws, row, 1, "Показатель", font=HEAD_FONT, fill=HEADER_FILL, align=WRAP_CENTER)
    put(ws, row, 2, "Закрытых случаев", font=HEAD_FONT, fill=HEADER_FILL, align=WRAP_CENTER)
    row += 1
    for indicator in ("6.1.3.2.8", "6.1.3.2.9", "6.1.3.2.10", "6.1.3.2.11"):
        put(ws, row, 1, "%s — %s" % (indicator, INDICATOR_TITLES[indicator]), align=WRAP)
        # SUMIF, а не SUMPRODUCT по диапазону месяцев: признак показателя —
        # одна колонка, месяцы — двенадцать, и перемножить их поэлементно
        # нельзя, размерности не совпадают. Суммируем колонку годового итога,
        # текст SUMIF из неё игнорирует сам.
        put(ws, row, 2,
            "=SUMIF('%s'!$F$5:$F$%d,\"%s\",'%s'!$%s$5:$%s$%d)"
            % (sheet, end, indicator, sheet, total, total, end),
            align=WRAP_CENTER, fmt="#,##0")
        row += 1

    # Красным — только то, что действительно ошибка: текст и отрицательные числа.
    red = PatternFill("solid", start_color="FFF4CCCC", end_color="FFF4CCCC")
    for target in ("B7", "B8"):
        ws.conditional_formatting.add(
            target, FormulaRule(formula=["%s>0" % target], fill=red))
    return ws


def main():
    organizations = read_organizations(SRC_DIR)
    assert len(organizations) == 37, len(organizations)

    wb = openpyxl.Workbook()
    howto = wb.active
    howto.title = "Как заполнять"
    build_howto(howto, organizations)

    data = wb.create_sheet("Закрытые случаи")
    layout = build_data_sheet(data, organizations, by_subdivision=False)

    detailed = wb.create_sheet("По подразделениям")
    build_data_sheet(detailed, organizations, by_subdivision=True)

    org_count, case_count = build_reference_sheets(wb, organizations)
    build_control(wb, layout)

    # Выпадающие списки: OID и код вида случая печатать руками незачем.
    wb.defined_names["MO_OIDS"] = DefinedName(
        "MO_OIDS", attr_text="'Справочник МО'!$A$5:$A$%d" % (4 + org_count))
    wb.defined_names["CASE_CODES"] = DefinedName(
        "CASE_CODES", attr_text="'Виды случая'!$A$5:$A$%d" % (4 + case_count))
    oid_dv = DataValidation(type="list", formula1="MO_OIDS", allow_blank=True,
                            showErrorMessage=False)
    case_dv = DataValidation(type="list", formula1="CASE_CODES", allow_blank=True,
                             showErrorMessage=False)
    detailed.add_data_validation(oid_dv)
    detailed.add_data_validation(case_dv)
    oid_dv.sqref = "A5:A404"
    case_dv.sqref = "G5:G404"

    wb.save(OUT)
    print("Сохранено:", OUT)
    print("Медорганизаций:", org_count, "· видов случая:", case_count,
          "· строк на листе «Закрытые случаи»:", org_count * case_count)
    print("Диапазон месяцев: %s..%s, итог в колонке %s"
          % (get_column_letter(layout["month_start"]),
             get_column_letter(layout["month_start"] + 11),
             get_column_letter(layout["total_col"])))


if __name__ == "__main__":
    main()
