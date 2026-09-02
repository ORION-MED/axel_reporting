"""Примеры срабатывания причин диагностики — форма для методолога.

Зачем. В файле согласования от 18.08.2026 методолог пометила три причины
«требует обсуждения» и попросила конкретику:

    № 14 subdivision_not_in_frmr — «мне надо посмотреть пример, какое
         подразделение дало такой результат»;
    № 16 organization_without_subdivisions — «при наличии конкретных примеров
         проблемы с расчетом, давайте рассмотрим их. Пока не понимаю,
         не встречала ранее такого»;
    № 18 requirement_waived_organization_absent_from_tpgg — «при наличии
         конкретных примеров давайте рассмотрим. Пока не понимаю».

Ответить на это можно только данными: причина срабатывает на реальных OID,
и пока их не видно, обсуждать нечего.

Лист на причину, плюс сводный лист с решением по каждой. Запросы к стенду —
в README.

Запуск:
    python build_examples.py subdiv.tsv waived.tsv out.xlsx
"""

from __future__ import annotations

import csv
import io
import sys
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DECISIONS = ['причина верна, оставить', 'причина не нужна, убрать', 'изменить текст']

HEADER_FILL = PatternFill('solid', fgColor='DDE5F0')
ANSWER_FILL = PatternFill('solid', fgColor='FFF6DA')
NOTE_FONT = Font(italic=True, color='555555')


def read_tsv(path: str, key: str) -> list[dict[str, str]]:
    """Строки выгрузки. `key` — колонка, без которой строка не данные.

    Проверка не формальная: psql дописывает в конец подвал «(89 rows)»,
    и без неё он становится ещё одной строкой — восемьдесят девять
    подразделений превращаются в девяносто, и цифра в файле расходится
    с той, что методолог видит на экране. Выгружать стоит
    с `--pset=footer=off`, но полагаться только на это нельзя.
    """
    with io.open(path, encoding='utf-8') as handle:
        return [row for row in csv.DictReader(handle, delimiter='\t') if row.get(key)]


def style_header(sheet, count: int) -> None:
    for column in range(1, count + 1):
        cell = sheet.cell(sheet.max_row, column)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def widths(sheet, values: list[int]) -> None:
    for index, width in enumerate(values, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_summary(sheet, subdiv: list[dict], waived: list[dict]) -> None:
    """Сводный лист — то, что читается первым и по чему принимается решение."""
    sheet.title = 'Три причины'
    sheet.append(['Примеры срабатывания причин, помеченных «требует обсуждения»'])
    sheet['A1'].font = Font(bold=True, size=13)
    sheet.append([
        'По каждой причине — отдельный лист с полным списком. '
        'Решение впишите в жёлтую колонку.'
    ])
    sheet.append([])

    sheet.append([
        '№', 'Код причины', 'Срабатываний',
        'Что означает на наших данных', 'Решение', 'Комментарий',
    ])
    style_header(sheet, 6)

    organizations = len({row['mo'] for row in waived if row.get('mo')})
    types = len({row['code'] for row in waived if row.get('code')})

    rows = [
        (
            14, 'subdivision_not_in_frmr', f'{len(subdiv)} подразделений',
            'Подразделения формируют документы, но их нет в ФРМР. Вид подразделения '
            'неизвестен, поэтому применимость по подразделению не считается. '
            'Крупнейшее — 22 798 документов у МРБ № 4.',
        ),
        (
            16, 'organization_without_subdivisions', '1 организация',
            'АО «Курганфармация»: в ФРМР нет ни одного её подразделения. '
            'Это аптечная организация, и подразделений медицинского профиля '
            'у неё в регистре действительно нет.',
        ),
        (
            18, 'requirement_waived_organization_absent_from_tpgg',
            f'{len(waived)} пар «МО × вид»',
            f'{types} вида у {organizations} МО. Все десять работают вне ОМС — '
            'психоневрологические, противотуберкулёзные, наркологический, СМЭ, '
            'патанатомия, станция переливания, аптека. Ровно тот случай, о котором '
            'Вы говорили: «госзадание шире, чем терпрограмма ОМС».',
        ),
    ]
    for number, code, count, meaning in rows:
        sheet.append([number, code, count, meaning, '', ''])
        sheet.cell(sheet.max_row, 5).fill = ANSWER_FILL
        sheet.cell(sheet.max_row, 6).fill = ANSWER_FILL
        sheet.cell(sheet.max_row, 4).alignment = Alignment(wrap_text=True, vertical='top')

    validation = DataValidation(
        type='list', formula1='"' + ','.join(DECISIONS) + '"', allow_blank=True,
    )
    sheet.add_data_validation(validation)
    validation.add(f'E{sheet.max_row - 2}:E{sheet.max_row}')

    widths(sheet, [6, 46, 18, 78, 26, 40])


def write_subdivisions(sheet, rows: list[dict]) -> None:
    sheet.title = '14 · нет в ФРМР'
    sheet.append(['Подразделения, формирующие документы, но отсутствующие в ФРМР'])
    sheet['A1'].font = Font(bold=True, size=13)
    sheet.append([
        f'Всего {len(rows)}. OID подразделения приходит в выгрузке РЭМД, '
        'но в ФРМР такого нет — вид подразделения определить не по чему.'
    ])
    sheet.append([
        'Практическое следствие: применимость видов по подразделению у этих строк '
        'не рассчитывается, документы при этом в числитель МО попадают.'
    ])
    sheet.cell(sheet.max_row, 1).font = NOTE_FONT
    sheet.append([])

    sheet.append(['№', 'Медицинская организация', 'OID подразделения', 'Документов'])
    style_header(sheet, 4)

    by_organization = Counter()
    for index, row in enumerate(rows, start=1):
        docs = int(row.get('docs') or 0)
        sheet.append([index, row.get('mo', ''), row.get('oid', ''), docs])
        by_organization[row.get('mo', '')] += docs

    sheet.append([])
    sheet.append(['Итого по медорганизациям'])
    sheet.cell(sheet.max_row, 1).font = Font(bold=True)
    for name, docs in by_organization.most_common():
        sheet.append(['', name, '', docs])

    widths(sheet, [6, 52, 46, 14])
    sheet.freeze_panes = 'A7'


def write_waived(sheet, rows: list[dict]) -> None:
    sheet.title = '18 · нет в ТПГГ'
    sheet.append(['Обязательность снята: медорганизации нет в терпрограмме ОМС'])
    sheet['A1'].font = Font(bold=True, size=13)
    sheet.append([
        'Основание правила — «наличие объёма МП, утверждённого госзаданием». '
        'Сервис проверяет его по файлу ТПГГ, другого источника объёмов у него нет.'
    ])
    sheet.append([
        'Поэтому МО, работающая вне ОМС, такое основание не проходит никогда — '
        'не потому, что помощь не оказывает, а потому, что её нет в терпрограмме.'
    ])
    sheet.cell(sheet.max_row, 1).font = NOTE_FONT
    sheet.append([])

    sheet.append(['Медицинская организация', 'Код вида', 'Вид СЭМД'])
    style_header(sheet, 3)

    by_organization: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_organization[row.get('mo', '')].append(row)

    for name in sorted(by_organization):
        for row in by_organization[name]:
            sheet.append([name, row.get('code', ''), row.get('type_name', '')])

    widths(sheet, [40, 12, 76])
    sheet.freeze_panes = 'A7'


def main() -> None:
    if len(sys.argv) != 4:
        print(__doc__)
        raise SystemExit(2)
    subdiv_path, waived_path, out_path = sys.argv[1:4]
    subdiv = read_tsv(subdiv_path, 'oid')
    waived = read_tsv(waived_path, 'code')

    workbook = Workbook()
    write_summary(workbook.active, subdiv, waived)
    write_subdivisions(workbook.create_sheet('14'), subdiv)
    write_waived(workbook.create_sheet('18'), waived)
    workbook.save(out_path)
    print(f'Готово: {out_path}; подразделений {len(subdiv)}, пар «МО × вид» {len(waived)}')


if __name__ == '__main__':
    main()
